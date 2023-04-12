import os
import re
import openpyxl
import requests
from bs4 import BeautifulSoup
import tkinter as tk
from tkinter import filedialog

# Define function to extract version number from URL
def extract_version(url, current_version):
    # Make HTTP request to URL
    response = requests.get(url)
    # Parse HTML content of webpage
    soup = BeautifulSoup(response.content, 'html.parser')
    # Extract version information using regular expressions
    version_pattern = re.compile(r'(\d+\.)?(\d+\.)?(\*|\d+)')
    match = version_pattern.search(str(soup))
    if match:
        extracted_version = match.group()
        if str(extracted_version) == str(current_version):
            return 'Same'
        else:
            return extracted_version
    else:
        return ''

# Define function to handle button click
def browse_excel_file():
    filename = filedialog.askopenfilename(initialdir='/', title='Select Excel file', filetypes=(('Excel Files', '*.xlsx'),))
    if filename:
        latest_file_name = extract_latest_version(filename)
        message_label.config(text=f'Latest file saved as {latest_file_name}')

# Define function to extract latest version and save new workbook
def extract_latest_version(filename):
    # Load workbook
    workbook = openpyxl.load_workbook(filename)

    # Find the next available number for the filename
    i = 1
    while True:
        latest_file_name = f'{os.path.splitext(filename)[0]}_Latest({i}).xlsx'
        if not os.path.isfile(latest_file_name):
            break
        i += 1

    # Create a copy of the sheet in a new workbook
    source_sheet = workbook['Sheet1']
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = 'Sheet1'

    # Add column headers
    new_sheet.cell(row=1, column=1, value='Webpage URL')
    new_sheet.cell(row=1, column=2, value='Version')
    new_sheet.cell(row=1, column=3, value='Latest Version')

    # Copy data from source sheet
    for row in source_sheet.iter_rows(min_row=2):
        new_row = []
        for cell in row:
            new_row.append(cell.value)
        version = new_row[1]
        url = new_row[0]
        extracted_version = extract_version(url, version)
        if extracted_version == "Error":
            new_row.append("Error")
        elif extracted_version == "Same":
            new_row.append("Same")
        else:
            new_row.append(extracted_version)
        new_sheet.append(new_row)

    # Save the new workbook
    new_workbook.save(latest_file_name)
    return latest_file_name

# Create GUI
root = tk.Tk()
root.title('Version Checker')
root.configure(bg='white')
root.geometry('600x300')

# Create label for instructions
instructions_label = tk.Label(root, text='Select an Excel file to check the versions:', font=('Arial', 12), bg='white')
instructions_label.pack(pady=20)

# Create browse button
browse_button = tk.Button(root, text='Browse', font=('Arial', 16), command=browse_excel_file)
browse_button.pack(pady=10)

# Create message label
message_label = tk.Label(root, font=('Arial', 12), bg='white')
message_label.pack(pady=20)

root.mainloop()